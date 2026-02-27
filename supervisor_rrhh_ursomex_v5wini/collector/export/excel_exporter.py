from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def _looks_like_placeholder_id(emp_id: Any) -> bool:
    s = str(emp_id).strip() if emp_id is not None else ""
    if not s:
        return False
    if not s.isdigit():
        return False
    # Unsigned placeholder (common when device sends -1 as uint64)
    if len(s) >= 18 and s.startswith("1844674407"):
        return True
    try:
        v = int(s)
        return v <= 0 or v >= 2**63
    except Exception:
        return False


def _display_emp_id(emp_id: Any, emp_name: Any) -> str:
    """Return a human-friendly employee_id for exports.

    If `employee_id` looks like a placeholder huge numeric, prefer the name (what you typed on the device).
    """
    if _looks_like_placeholder_id(emp_id):
        n = str(emp_name).strip() if emp_name is not None else ""
        return n or str(emp_id)
    return str(emp_id).strip() if emp_id is not None else ""


def export_excel_jornadas_summary(
    jornadas_rows: List[Dict[str, Any]],
    out_dir: str,
    name: str,
    template_path: str = "",
    extra_sheets: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    export_id: str = "",
    meta: Optional[Dict[str, Any]] = None,
) -> str:

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    outfile = os.path.join(out_dir, f"{name}.xlsx")

    wb: Workbook
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
        if "JORNADAS_CIERRE" not in wb.sheetnames:
            wb.create_sheet("JORNADAS_CIERRE")
        ws = wb["JORNADAS_CIERRE"]
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "JORNADAS_CIERRE"


    # Keep public columns stable, then append technical metadata columns at the end.
    headers = [
        "fecha_registro",
        "employee_id",
        "employee_name",
        "start_local",
        "end_local",
        "duration_minutes",
    ] + [f"E{i:02d}" for i in range(1, 13)] + [
        "incidencia_codes",
        "incidencia_detail",
        # Technical columns (used by CORRECCIONES importer). Keep at the end.
        "jornada_uid",
        "__jornada_id",
        "__start_time_utc",
    ]
    ws.append(headers)

    for r in jornadas_rows:
        row = []
        for h in headers:
            if h == "employee_id":
                row.append(_display_emp_id(r.get("employee_id"), r.get("employee_name")))
            elif h == "__jornada_id":
                row.append(r.get("jornada_id", ""))
            elif h == "__start_time_utc":
                row.append(r.get("start_time_utc", ""))
            else:
                row.append(r.get(h, ""))
        ws.append(row)

    # Hide technical columns (but keep jornada_uid visible so RRHH can copy/paste into CORRECCIONES)
    try:
        # Find column letters by header name
        header_row = [c.value for c in ws[1]]
        for key in ("__jornada_id", "__start_time_utc"):
            if key in header_row:
                idx = header_row.index(key) + 1
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].hidden = True
    except Exception:
        pass

    # Meta sheet (export_id + range) so the importer can rebuild without guessing.
    if not export_id:
        # deterministic enough for a single export; persisted in the file
        export_id = datetime.utcnow().strftime("EXPORT_%Y%m%d_%H%M%S")
    meta = meta or {}
    if "__META" not in wb.sheetnames:
        wb.create_sheet("__META")
    ws_meta = wb["__META"]
    ws_meta.delete_rows(1, ws_meta.max_row)
    ws_meta.append(["export_id", export_id])
    ws_meta.append(["generated_at_utc", datetime.utcnow().replace(microsecond=0).isoformat() + "Z"])
    for k in ["range_start_op", "range_end_op", "source_label", "file_name", "app_version"]:
        v = meta.get(k, "") if isinstance(meta, dict) else ""
        if v:
            ws_meta.append([k, v])
    # hide meta sheet
    try:
        ws_meta.sheet_state = "hidden"
    except Exception:
        pass

    # CORRECCIONES sheet (manual overrides for jornada closure)
    if "CORRECCIONES" not in wb.sheetnames:
        wb.create_sheet("CORRECCIONES")
    ws_corr = wb["CORRECCIONES"]
    ws_corr.delete_rows(1, ws_corr.max_row)
    ws_corr.append(["CORRECCIONES (Jornada)"])
    ws_corr.append(["", "Usa esta hoja para forzar la interpretación de cierre cuando un evento del día siguiente debería pegar al día anterior (o viceversa)."])
    ws_corr.append(["", "No edites JORNADAS_CIERRE; aquí solo agregas filas."])
    ws_corr.append([])
    ws_corr.append(["jornada_uid", "accion", "nota"])
    # data validation for accion
    dv = DataValidation(type="list", formula1='"FORZAR_CIERRE_D1,FORZAR_ENTRADA"', allow_blank=True)
    ws_corr.add_data_validation(dv)
    dv.add("B6:B500")
    try:
        ws_corr.freeze_panes = "A6"
    except Exception:
        pass

    # Optional extra sheets (e.g., model audit)
    if extra_sheets and isinstance(extra_sheets, dict):
        for sheet_name, data in extra_sheets.items():
            if not sheet_name:
                continue
            sname = str(sheet_name)[:31]
            if sname not in wb.sheetnames:
                wb.create_sheet(sname)
            wsx = wb[sname]
            wsx.delete_rows(1, wsx.max_row)
            data = data or []
            if not data:
                wsx.append(["(sin registros)"])
                continue
            # Keep header order stable: common audit columns first, then any extras.
            preferred = [
                "op_date",
                "employee_id",
                "event_time_local",
                "decision",
                "confidence",
                "p_prior",
                "reasons",
                "boundary_to_op_date",
                "seq_before",
                "min_prev",
                "allow_no_late",
                "has_late_signature",
                "lookahead_cnt",
                "created_at_utc",
            ]
            keys = set()
            for d in data:
                if isinstance(d, dict):
                    keys |= set(d.keys())
            headers_x = [k for k in preferred if k in keys] + [k for k in sorted(keys) if k not in preferred]
            wsx.append(headers_x)
            for d in data:
                if not isinstance(d, dict):
                    continue
                wsx.append([d.get(h, "") for h in headers_x])

    wb.save(outfile)
    return outfile
