from __future__ import annotations

import os
from contextlib import nullcontext
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from collector.storage.db import DB


ALLOWED_DECISIONS = {
    "FORZAR_CIERRE_D1",  # tratar el primer evento del día siguiente como OUT del día anterior
    "FORZAR_ENTRADA",    # NO permitir carryback; el evento del día siguiente es IN de nueva jornada
}


@dataclass
class ImportResult:
    export_id: str
    imported: int
    skipped: int
    errors: int
    decisions: Dict[str, int]


def _read_meta_export_id(wb) -> str:
    try:
        if "__META" not in wb.sheetnames:
            return ""
        ws = wb["__META"]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
            if not row or len(row) < 2:
                continue
            k = str(row[0] or "").strip()
            if k == "export_id":
                return str(row[1] or "").strip()
    except Exception:
        return ""
    return ""


def import_manual_corrections_excel(db: DB, excel_path: str) -> ImportResult:
    """Import manual jornada corrections from an exported Excel.

    Expected workbook structure:
      - __META sheet with export_id
      - CORRECCIONES sheet with header row: jornada_uid | accion | nota

    Behavior:
      - upsert_manual_label(jornada_uid, accion)
      - insert_manual_correction(export_id, jornada_uid, accion, nota)
    """

    if not excel_path or not os.path.exists(excel_path):
        raise FileNotFoundError(f"No existe el Excel: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    export_id = _read_meta_export_id(wb)
    if not export_id:
        export_id = "UNKNOWN_EXPORT"

    if "CORRECCIONES" not in wb.sheetnames:
        raise ValueError("El Excel no tiene la hoja CORRECCIONES")
    ws = wb["CORRECCIONES"]

    # Find header row that contains jornada_uid/accion
    header_row_idx: Optional[int] = None
    headers: List[str] = []
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [str((c.value or "")).strip() for c in ws[r]]
        low = [v.lower() for v in vals]
        if "jornada_uid" in low and ("accion" in low or "acción" in low):
            header_row_idx = r
            headers = low
            break

    if header_row_idx is None:
        raise ValueError("No se encontró encabezado válido en CORRECCIONES (requiere jornada_uid y accion)")

    def _col(name: str) -> Optional[int]:
        n = name.lower()
        if n in headers:
            return headers.index(n)
        # allow spanish accent
        if n == "accion" and "acción" in headers:
            return headers.index("acción")
        return None

    c_uid = _col("jornada_uid")
    c_acc = _col("accion")
    c_note = _col("nota")
    if c_uid is None or c_acc is None:
        raise ValueError("Encabezado incompleto en CORRECCIONES")

    imported = 0
    skipped = 0
    errors = 0
    decisions: Dict[str, int] = {}

    tx = db.transaction() if getattr(db, 'transaction', None) and db.engine == 'sqlite' else nullcontext()
    with tx:
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
            try:
                if not row:
                    continue
                uid = str(row[c_uid] or "").strip()
                acc = str(row[c_acc] or "").strip().upper()
                note = ""
                if c_note is not None and c_note < len(row):
                    note = str(row[c_note] or "").strip()
                if not uid and not acc and not note:
                    continue
                if not uid or not acc:
                    skipped += 1
                    continue
                if acc not in ALLOWED_DECISIONS:
                    skipped += 1
                    continue

                db.upsert_manual_label(uid, acc, note=note)
                db.insert_manual_correction(export_id, uid, acc, note=note)
                decisions[acc] = decisions.get(acc, 0) + 1
                imported += 1
            except Exception:
                errors += 1

    return ImportResult(export_id=export_id, imported=imported, skipped=skipped, errors=errors, decisions=decisions)
