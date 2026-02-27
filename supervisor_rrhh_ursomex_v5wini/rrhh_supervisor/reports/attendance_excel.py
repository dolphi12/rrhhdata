from __future__ import annotations

from datetime import datetime, timedelta
from typing import Any, Dict, List
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from rrhh_supervisor.storage.db import DB
from rrhh_supervisor.reports.i18n_es import parse_iso_date, date_es, weekday_es
from rrhh_supervisor.reports.motivation import quote_of_day
from rrhh_supervisor.services.worktime import build_events_by_jornada_id, compute_net_minutes_from_events



def _parse_utc(iso: str) -> Any:
    s = (iso or "").strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(ZoneInfo("UTC"))
    except Exception:
        return None


def _to_local_hhmm(iso_utc: str, local_tz: str) -> str:
    dt_utc = _parse_utc(iso_utc)
    if dt_utc is None:
        return ""
    dt_local = dt_utc.astimezone(ZoneInfo(local_tz))
    return f"{dt_local.hour:02d}:{dt_local.minute:02d}"


def _to_local_date(iso_utc: str, local_tz: str) -> str:
    dt_utc = _parse_utc(iso_utc)
    if dt_utc is None:
        return ""
    return dt_utc.astimezone(ZoneInfo(local_tz)).date().isoformat()


def _style_header(ws, last_col_letter: str):
    fill = PatternFill("solid", fgColor="008C8C")  # teal
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def _style_grid(ws, start_row: int, start_col: int, end_row: int, end_col: int):
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
def _utc_iso(dt_utc: datetime) -> str:
    return dt_utc.replace(microsecond=0).isoformat().replace("+00:00", "Z")


def export_attendance_matrix(collector: DB, store: DB,
    local_tz: str,
    start_date: str,
    end_date: str,
    out_path: str,
) -> str:
    tz = ZoneInfo(local_tz)
    start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=tz)
    end = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=tz) + timedelta(days=1)
    start_utc = start.astimezone(ZoneInfo("UTC"))
    end_utc = end.astimezone(ZoneInfo("UTC"))
    jornadas = collector.list_jornadas_closed_range(None, _utc_iso(start_utc), _utc_iso(end_utc))

    dates = []
    d = start.date()
    while d < end.date():
        dates.append(d.isoformat())
        d = d + timedelta(days=1)

    # Permisos (días no laborables justificados)
    perms_map = store.permissions_set_opdate_range(start_date, end_date)

    date_labels = []
    for s in dates:
        dd = parse_iso_date(s)
        if dd:
            date_labels.append(f"{weekday_es(dd, abbr=True)} {date_es(dd)}")
        else:
            date_labels.append(s)

    roster = store.list_roster(active_only=True)


    # --- Cálculo RRHH (neto) consistente ---
    emp_bounds: Dict[str, Dict[str, str]] = {}
    for jrow in jornadas:
        eid = str(jrow.get("employee_id") or "").strip()
        if not eid:
            continue
        su = str(jrow.get("start_time_utc") or "")
        eu = str(jrow.get("end_time_utc") or "")
        b = emp_bounds.get(eid)
        if b is None:
            b = {"min": su or "", "max": eu or ""}
            emp_bounds[eid] = b
        if su and (not b["min"] or su < b["min"]):
            b["min"] = su
        if eu and (not b["max"] or eu > b["max"]):
            b["max"] = eu

    events_by_jid_all: Dict[str, List[Dict[str, Any]]] = {}
    for eid, b in emp_bounds.items():
        if not b.get("min") or not b.get("max"):
            continue
        try:
            events_by_jid_all.update(build_events_by_jornada_id(collector, eid, b["min"], b["max"]))
        except Exception:
            continue

    net_minutes_by_jid: Dict[str, int] = {}
    for jrow in jornadas:
        jid = str(jrow.get("jornada_id") or "").strip()
        if not jid:
            continue
        calc = compute_net_minutes_from_events(events_by_jid_all.get(jid, []))
        nm = int(calc.net_minutes or 0)
        if nm <= 0:
            nm = int(jrow.get("duration_minutes") or 0)
        net_minutes_by_jid[jid] = nm

    emp_map: Dict[str, Dict[str, Any]] = {}
    if roster:
        for r in roster:
            emp = str(r.get("employee_id") or "").strip()
            if not emp:
                continue
            emp_map.setdefault(
                emp,
                {"employee_name": str(r.get("employee_name") or "").strip(), "days": {}, "hours": {}},
            )

    for j in jornadas:
        emp = str(j.get("employee_id") or "").strip()
        if not emp:
            continue
        emp_map.setdefault(emp, {"employee_name": str(j.get("employee_name") or "").strip(), "days": {}, "hours": {}})
        od = str(j.get("op_date") or "").strip()
        if od:
            emp_map[emp]["days"][od] = True
            jid = str(j.get("jornada_id") or "").strip()
            emp_map[emp]["hours"][od] = float(net_minutes_by_jid.get(jid, int(j.get("duration_minutes") or 0))) / 60.0
            # Marca cierre en D+1 (fin de jornada al día siguiente)
            end_iso = str(j.get("end_time_utc") or "")
            cross = False
            if end_iso and od:
                ld = _to_local_date(end_iso, local_tz)
                if ld and ld > od:
                    cross = True
            emp_map[emp].setdefault("cross", {})
            emp_map[emp]["cross"][od] = cross
            # Notas (si existen en la jornada)
            note_raw = j.get("incidencia_codes") if isinstance(j.get("incidencia_codes"), (str, list)) else j.get("notes")
            if isinstance(note_raw, list):
                note_raw = ",".join([str(x) for x in note_raw])
            emp_map[emp].setdefault("notes", {})
            emp_map[emp]["notes"][od] = str(note_raw or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    headers = ["ID", "Nombre"] + date_labels
    ws.append(headers)

    for emp in sorted(emp_map.keys()):
        row = [emp, emp_map[emp]["employee_name"]]
        for od in dates:
            if emp_map[emp]["days"].get(od):
                h = emp_map[emp]["hours"].get(od, 0)
                cross = bool(emp_map[emp].get("cross", {}).get(od))
                hh = float(h or 0.0)
                cell = f"P {hh:.1f}"
                if cross:
                    cell += "*"
                row.append(cell)
            else:
                # Permiso -> no cuenta como ausencia
                if od in perms_map.get(emp, set()):
                    row.append("PR")
                else:
                    row.append("A")
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col == 1 else (26 if col == 2 else 13)


    # Estilos (más 'profesional')
    last_col = get_column_letter(len(headers))
    _style_header(ws, last_col)
    # Alineaciones
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        for c in range(3, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    _style_grid(ws, 1, 1, ws.max_row, len(headers))

    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Rango", f"{start_date} a {end_date}"])
    ws2.append(["Nota semana operativa", "La semana operativa va de Miércoles a Martes (corte día operativo 03:00)."])
    ws2.append(["Empleados considerados", len(emp_map)])
    ws2.append(["Leyenda", "P = Presente (horas registradas), A = Ausente o sin jornada cerrada, PR = Permiso (no cuenta como ausencia)"])
    ws2.append(["Frase del día", quote_of_day(local_tz)])

    
    # Hoja Detalle (similar al PDF: registro diario)
    ws3 = wb.create_sheet("Detalle")
    ws3.append(["ID", "Nombre", "Fecha (día operativo)", "Primera", "Última", "Dur (h)", "D+1", "Notas"])
    # Construimos un índice por (emp, op_date) desde jornadas
    # NOTA: aquí no listamos intermedios para mantenerlo simple/legible.
    for j in jornadas:
        emp = str(j.get("employee_id") or "").strip()
        if not emp:
            continue
        name = str(j.get("employee_name") or emp_map.get(emp, {}).get("employee_name", "")).strip()
        od = str(j.get("op_date") or "").strip()
        st = _to_local_hhmm(str(j.get("start_time_utc") or ""), local_tz)
        en = _to_local_hhmm(str(j.get("end_time_utc") or ""), local_tz)
        hh = float(j.get("duration_minutes") or 0.0) / 60.0
        cross = False
        if od:
            ld = _to_local_date(str(j.get("end_time_utc") or ""), local_tz)
            if ld and ld > od:
                cross = True
        note_raw = j.get("incidencia_codes") if isinstance(j.get("incidencia_codes"), (str, list)) else j.get("notes")
        if isinstance(note_raw, list):
            note_raw = ",".join([str(x) for x in note_raw])
        note = str(note_raw or "").strip()
        ws3.append([emp, name, od, st, en, round(hh, 1), "Sí" if cross else "No", note])

    # Estilos Detalle
    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = [10, 28, 18, 10, 10, 10, 7, 28][col - 1]
    _style_header(ws3, get_column_letter(8))
    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws3.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws3.cell(row=r, column=8).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for c in range(3, 8):
            ws3.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    _style_grid(ws3, 1, 1, ws3.max_row, 8)

    # Resumen con notas más completas (similar al PDF)
    ws2.append([])
    ws2.append(["Notas y convenciones", ""])
    ws2.append(["P", "Presente (jornada cerrada en el día operativo)"])
    ws2.append(["A", "Ausente o sin jornada cerrada"])
    ws2.append(["*", "Cierre registrado en D+1 (fin de jornada al día siguiente)"])
    ws2.append(["Semana operativa", "Miércoles a Martes (corte día operativo 03:00)"])
    wb.save(out_path)
    return out_path
